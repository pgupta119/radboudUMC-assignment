import xlrd
import datetime
from openpyxl.workbook import Workbook
import os
import logging
import openpyxl
# Directories for source .xls files and output .xlsx files
xls_dir = '/Users/prakashgupta/radboudUMC-assignment/data/source/'
xlsx_dir = '/data/output/'
# Setup logging with both file and console handlers
def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    file_handler = logging.FileHandler('conversion_log.log')
    console_handler = logging.StreamHandler()

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

setup_logging()

import datetime


def convert_xls_to_xlsx(src_file_path, dest_file_path):
    try:
        xls_book = xlrd.open_workbook(src_file_path)
        xlsx_book = Workbook()

        for sheet_index, sheet_name in enumerate(xls_book.sheet_names()):
            xls_sheet = xls_book.sheet_by_index(sheet_index)
            xlsx_sheet = xlsx_book.active if sheet_index == 0 else xlsx_book.create_sheet(title=sheet_name)

            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row, col)
                    cell_type = xls_sheet.cell_type(row, col)

                    # Check if cell type is date
                    if cell_type == xlrd.XL_CELL_DATE:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                        cell_value = datetime.datetime(*date_tuple)

                    xlsx_sheet.cell(row=row + 1, column=col + 1).value = cell_value

        xlsx_book.save(dest_file_path)
        return True
    except Exception as e:
        logging.error(f"Error converting file {src_file_path}: {e}")
        return False


def verify_conversion(src_file_path, dest_file_path):
    try:
        xls_book = xlrd.open_workbook(src_file_path)
        xlsx_book = openpyxl.load_workbook(dest_file_path)

        for sheet_name in xls_book.sheet_names():
            if sheet_name not in xlsx_book.sheetnames:
                return False

            xls_sheet = xls_book.sheet_by_name(sheet_name)
            xlsx_sheet = xlsx_book[sheet_name]

            for row in range(1, xls_sheet.nrows + 1):
                for col in range(1, xls_sheet.ncols + 1):
                    if xls_sheet.cell_value(row - 1, col - 1) != xlsx_sheet.cell(row, col).value:
                        return False

        return True
    except Exception as e:
        logging.error(f"Error verifying file {src_file_path}: {e}")
        return False

def batch_process(src_directory, dest_directory, batch_size=5):
    for subdir, _, files in sorted(os.walk(src_directory)):
        relative_path = os.path.relpath(subdir, src_directory)
        dest_subdir = os.path.join(dest_directory, relative_path)
        os.makedirs(dest_subdir, exist_ok=True)

        xls_files = sorted([f for f in files if f.endswith('.xls')])
        total_files = len(xls_files)
        for i in range(0, total_files, batch_size):
            batch = xls_files[i:i + batch_size]
            for file in batch:
                src_file_path = os.path.join(subdir, file)
                dest_file_name = os.path.splitext(file)[0] + '.xlsx'
                dest_file_path = os.path.join(dest_subdir, dest_file_name)

                # Check if the file has already been converted
                if not os.path.exists(dest_file_path):
                    if convert_xls_to_xlsx(src_file_path, dest_file_path):
                        if verify_conversion(src_file_path, dest_file_path):
                            logging.info(f"Successfully converted and verified {file}")
                        else:
                            logging.error(f"Verification failed for {file}")
                    else:
                        logging.error(f"Conversion failed for {file}")
                else:
                    logging.info(f"File already converted: {file}")




batch_process(xls_dir, xlsx_dir)
